"""Reproducibly export the reviewed v11.5 workbook into the web snapshot.

Formula text and the last values cached by Excel are preserved separately. The
web app never recalculates the forensic source workbook.
"""

from datetime import datetime, timezone
from pathlib import Path
import hashlib
import json

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
INPUT = ROOT / "plan" / "UKD_Grow_Masterplan_2026_v11_5_CLOCKS_PH_BALANCE_MIXED_AUTO_3x9L.xlsx"
OUTPUT = ROOT / "public" / "data" / "evidence-guarded-workbook-v11_5.json"


class Encoder(json.JSONEncoder):
    def default(self, value):
        if isinstance(value, datetime):
            normalized = value if value.tzinfo else value.replace(tzinfo=timezone.utc)
            return normalized.astimezone(timezone.utc).isoformat().replace("+00:00", "Z")
        return super().default(value)


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def main() -> None:
    if not INPUT.exists():
        raise SystemExit(f"Missing source workbook: {INPUT}")

    formulas_book = load_workbook(INPUT, data_only=False, read_only=False)
    values_book = load_workbook(INPUT, data_only=True, read_only=False)
    payload: dict[str, dict[str, object]] = {}

    for sheet_name in formulas_book.sheetnames:
        formula_sheet = formulas_book[sheet_name]
        value_sheet = values_book[sheet_name]
        max_row = formula_sheet.max_row or 1
        max_column = formula_sheet.max_column or 1
        values: list[list[object]] = []
        formulas: list[list[object]] = []
        for row_number in range(1, max_row + 1):
            value_row: list[object] = []
            formula_row: list[object] = []
            for column_number in range(1, max_column + 1):
                formula_cell = formula_sheet.cell(row_number, column_number)
                value_cell = value_sheet.cell(row_number, column_number)
                if formula_cell.data_type == "f":
                    formula_row.append(formula_cell.value)
                    value_row.append(value_cell.value)
                else:
                    formula_row.append(None)
                    value_row.append(formula_cell.value)
            values.append(value_row)
            formulas.append(formula_row)
        payload[sheet_name] = {
            "range": formula_sheet.calculate_dimension(),
            "values": values,
            "formulas": formulas,
        }

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(
        json.dumps(payload, ensure_ascii=False, cls=Encoder, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"sheets={len(payload)}")
    print(f"source_sha256={sha256(INPUT)}")
    print(f"snapshot_sha256={sha256(OUTPUT)}")


if __name__ == "__main__":
    main()
