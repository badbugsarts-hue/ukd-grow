import openpyxl
import json
import sys
import os
import hashlib
from datetime import datetime

class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)

def get_file_sha256(filepath):
    h = hashlib.sha256()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(4096), b""):
            h.update(chunk)
    return h.hexdigest().upper()

def main():
    root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    input_file = os.path.join(root_dir, 'UKD_Grow_Masterplan_Elite_2026_v8_OWNED_STOCK_HESI_AN.xlsx')
    output_file = os.path.join(root_dir, 'public', 'data', 'evidence-guarded-workbook-v8.json')

    print(f"Loading workbook: {input_file}")
    wb = openpyxl.load_workbook(input_file, data_only=False)
    
    output = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        values = []
        formulas = []
        
        for row in ws.iter_rows():
            row_values = []
            row_formulas = []
            for cell in row:
                if cell.data_type == 'f':
                    # It's a formula.
                    row_formulas.append(cell.value)
                    row_values.append(None) 
                else:
                    row_formulas.append(None)
                    row_values.append(cell.value)
            values.append(row_values)
            formulas.append(row_formulas)
            
        output[sheet_name] = {
            "range": ws.dimensions,
            "values": values,
            "formulas": formulas
        }

    # Now load with data_only=True to get the evaluated values
    print("Loading workbook with data_only=True to extract evaluated values...")
    wb_data = openpyxl.load_workbook(input_file, data_only=True)
    for sheet_name in wb_data.sheetnames:
        ws_data = wb_data[sheet_name]
        
        for r_idx, row in enumerate(ws_data.iter_rows()):
            for c_idx, cell in enumerate(row):
                if output[sheet_name]["formulas"][r_idx][c_idx] is not None:
                    # It's a formula, replace the None in values with the evaluated value
                    output[sheet_name]["values"][r_idx][c_idx] = cell.value

    # Convert to JSON
    print(f"Writing JSON to: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, cls=DateTimeEncoder)

    print("Hashing files...")
    xlsx_hash = get_file_sha256(input_file)
    json_hash = get_file_sha256(output_file)

    print(f"XLSX SHA256: {xlsx_hash}")
    print(f"JSON SHA256: {json_hash}")
    print("Done!")

if __name__ == '__main__':
    main()
